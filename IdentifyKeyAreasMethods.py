import networkx as nx
import openpyxl as xl
import numpy as np
import copy
import matplotlib.pyplot as plt
from CentralityMenthod import *
import pandas as pd
import builtins
from openpyxl import load_workbook
import pandas as pd
import time
import os

def load_network(file):
    wb = xl.load_workbook(file)
    ws_node, ws_link = wb['node'], wb['link']
    row_link, row_node = ws_link.max_row, ws_node.max_row
    edge_list = list(ws_link.iter_rows(min_col=1, max_col=2, min_row=2, max_row=row_link, values_only=True))
    node_list = list(ws_node.iter_rows(min_col=1, max_col=3, min_row=2, max_row=row_node, values_only=True))
    G = nx.Graph()
    for i in node_list:
        G.add_node(i[0])
    for i in edge_list:
        G.add_edge(i[0], i[1])

    x_all, y_all = [i[1] for i in node_list], [i[2] for i in node_list]
    xcor_min, xcor_max = min(x_all), max(x_all)
    xcor = xcor_max - xcor_min
    ycor_min, ycor_max = min(y_all), max(y_all)
    ycor = ycor_max - ycor_min
    node_position = {}  # 节点的归一化位置坐标
    for i in node_list:
        node_position[i[0]] = list(
            (np.array([i[1], i[2]]) - np.array([xcor_min, ycor_min])) / np.array([xcor, ycor]))  # 计算每个节点的归一化位置

    return G, node_position



def PerformanceEvaluation(net):
    """计算最大连通子图比例和自然连通度"""
    size = len(nx.nodes(net))
    largest_cc = max(nx.connected_components(net), key=len)  # 获取最大的连通子图
    largest_cc = len(largest_cc) / size

    return largest_cc



def CreateGrid(interval):
    """生成网格字典,键为网格索引，值为网格四个点的坐标"""
    part = int(1 / interval)  # 生成一个 part × part 的二维网格
    grid = [[[(round(interval * m, 4), round(interval * n, 4)),
              (round(interval * (m + 1), 4), round(interval * n, 4)),
              (round(interval * m, 4), round(interval * (n + 1), 4)),
              (round(interval * (m + 1), 4), round(interval * (n + 1), 4))]
              for n in range(part)] for m in range(part)]

    # 将二维网格转换为一维列表
    Grid_flat = [coord for sublist in grid for coord in sublist]
    # 给每个网格单元添加一个索引
    Grid = {i: Grid_flat[i] for i in range(len(Grid_flat))}

    return Grid
#


epsilon = 1e-9  # 定义一个小的epsilon值
# 比较两个浮点数是否在epsilon范围内相等
def is_close(a, b, epsilon):
    return abs(a - b) < epsilon

def FindEdgesInGrid(net, Grid, pos):
    edge_list = list(net.edges)
    cen_edge, weight = {}, {}
    grid_index = list(Grid.keys())
    # print("网格索引：", grid_index)
    grid_pos = list(Grid.values())
    # print("网格四个点坐标：", grid_pos)
    for i in grid_index:
        weight[i] = 0
        cen_edge[i] = []
        for edge in edge_list:
            cor = [pos[edge[0]], pos[edge[1]]]  # 判断至少有一个端点在方格内
            if (grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]) or \
               (grid_pos[i][0][0] <= cor[1][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[1][1] <= grid_pos[i][2][1]):
                cen_edge[i].append(edge)
                weight[i] += 1
            else:  # 两个端点都不在方格内
                if cor[0][0] - cor[1][0] == 0:  # 平行y轴

                    if grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0]:
                        # 检查网格单元格的y坐标是否在线段的y坐标范围内
                        miny = min(cor[0][1], cor[1][1])
                        maxy = max(cor[0][1], cor[1][1])
                        if miny <= grid_pos[i][0][1] and maxy >= grid_pos[i][2][1]:
                            cen_edge[i].append(edge)
                            weight[i] += 1



                elif cor[0][1] - cor[1][1] == 0:  # 平行x轴

                    if grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]:
                        # 检查网格单元格的y坐标是否在线段的y坐标范围内
                        minx = min(cor[0][0], cor[1][0])
                        maxx = max(cor[0][0], cor[1][0])
                        if minx <= grid_pos[i][0][0] and maxx >= grid_pos[i][1][0]:
                            cen_edge[i].append(edge)
                            weight[i] += 1

                else:

                    slope = (cor[1][1] - cor[0][1]) / (cor[1][0] - cor[0][0])  # 斜率
                    intercept = cor[0][1] - slope * cor[0][0]  # 截距
                    # print(slope)
                    # print(intercept)

                    ver_x1 = grid_pos[i][0][0]
                    ver_y1 = slope * ver_x1 + intercept

                    ver_x2 = grid_pos[i][1][0]
                    ver_y2 = slope * ver_x2 + intercept

                    ver_y3 = grid_pos[i][0][1]
                    ver_x3 = (ver_y3 - intercept)/slope

                    ver_y4 = grid_pos[i][2][1]
                    ver_x4 = (ver_y4 - intercept)/slope


                    min_x_cor = min(cor[0][0], cor[1][0])
                    max_x_cor = max(cor[0][0], cor[1][0])
                    min_y_cor = min(cor[0][1], cor[1][1])
                    max_y_cor = max(cor[0][1], cor[1][1])

                    min_x_squa = min(grid_pos[i][0][0], grid_pos[i][1][0])
                    max_x_squa = max(grid_pos[i][0][0], grid_pos[i][1][0])
                    min_y_squa = min(grid_pos[i][0][1], grid_pos[i][2][1])
                    max_y_squa = max(grid_pos[i][0][1], grid_pos[i][2][1])

                    # 使用 is_close 函数来比较浮点数
                    points = [(ver_x1, ver_y1), (ver_x2, ver_y2), (ver_x3, ver_y3), (ver_x4, ver_y4)]
                    for point in points:
                        if is_close(point[0], min_x_cor, epsilon) or is_close(point[0], max_x_cor, epsilon) or (min_x_cor < point[0] < max_x_cor):
                            if is_close(point[1], min_y_cor, epsilon) or is_close(point[1], max_y_cor, epsilon) or (min_y_cor < point[1] < max_y_cor):
                                if is_close(point[0], min_x_squa, epsilon) or is_close(point[0], max_x_squa,epsilon) or (min_x_squa < point[0] < max_x_squa):
                                    if is_close(point[1], min_y_squa, epsilon) or is_close(point[1], max_y_squa,epsilon) or (min_y_squa < point[1] < max_y_squa):
                                        cen_edge[i].append(edge)
                                        weight[i] += 1
                                        break

    return cen_edge, weight



def AverageDegreeAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)

    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += net.degree(j)

    for i in grid_index:
        if len(cen_node[i]) > 0:
            weight[i] = weight[i] / len(cen_node[i])
        else:
            weight[i] = 0

    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print("DC: 删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:                  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))      # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net



def DegreeAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)

    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += net.degree(j)
    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print("DC: 删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:                  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))      # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net


def BetweennessAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_dict = nx.betweenness_centrality(net)
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)

    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += cen_dict[j]
    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print(max_cen_index)
    # print("删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))  # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net



def EigenvectorAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_dict = nx.eigenvector_centrality(net, tol=1.0e-3)
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)
    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += cen_dict[j]
    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print("EC: 删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))  # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net



def ClosenessAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_dict = nx.closeness_centrality(net)
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)

    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += cen_dict[j]
    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print("删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))  # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net


def NumEdges(net, Grid, pos, num):
    edge_list = list(net.edges)
    cen_edge, weight = {}, {}
    grid_index = list(Grid.keys())
    # print("网格索引：", grid_index)
    grid_pos = list(Grid.values())
    # print("网格四个点坐标：", grid_pos)
    for i in grid_index:
        weight[i] = 0
        cen_edge[i] = []
        for edge in edge_list:
            cor = [pos[edge[0]], pos[edge[1]]]  # 判断至少有一个端点在方格内
            if (grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[0][1] <=
                grid_pos[i][2][1]) or \
                    (grid_pos[i][0][0] <= cor[1][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[1][1] <=
                     grid_pos[i][2][1]):
                cen_edge[i].append(edge)
                weight[i] += 1
            else:  # 两个端点都不在方格内
                if cor[0][0] - cor[1][0] == 0:  # 平行y轴

                    if grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0]:
                        # 检查网格单元格的y坐标是否在线段的y坐标范围内
                        miny = min(cor[0][1], cor[1][1])
                        maxy = max(cor[0][1], cor[1][1])
                        if miny <= grid_pos[i][0][1] and maxy >= grid_pos[i][2][1]:
                            cen_edge[i].append(edge)
                            weight[i] += 1



                elif cor[0][1] - cor[1][1] == 0:  # 平行x轴

                    if grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]:
                        # 检查网格单元格的y坐标是否在线段的y坐标范围内
                        minx = min(cor[0][0], cor[1][0])
                        maxx = max(cor[0][0], cor[1][0])
                        if minx <= grid_pos[i][0][0] and maxx >= grid_pos[i][1][0]:
                            cen_edge[i].append(edge)
                            weight[i] += 1

                else:

                    slope = (cor[1][1] - cor[0][1]) / (cor[1][0] - cor[0][0])  # 斜率
                    intercept = cor[0][1] - slope * cor[0][0]  # 截距
                    # print(slope)
                    # print(intercept)

                    ver_x1 = grid_pos[i][0][0]
                    ver_y1 = slope * ver_x1 + intercept

                    ver_x2 = grid_pos[i][1][0]
                    ver_y2 = slope * ver_x2 + intercept

                    ver_y3 = grid_pos[i][0][1]
                    ver_x3 = (ver_y3 - intercept) / slope

                    ver_y4 = grid_pos[i][2][1]
                    ver_x4 = (ver_y4 - intercept) / slope

                    min_x_cor = min(cor[0][0], cor[1][0])
                    max_x_cor = max(cor[0][0], cor[1][0])
                    min_y_cor = min(cor[0][1], cor[1][1])
                    max_y_cor = max(cor[0][1], cor[1][1])

                    min_x_squa = min(grid_pos[i][0][0], grid_pos[i][1][0])
                    max_x_squa = max(grid_pos[i][0][0], grid_pos[i][1][0])
                    min_y_squa = min(grid_pos[i][0][1], grid_pos[i][2][1])
                    max_y_squa = max(grid_pos[i][0][1], grid_pos[i][2][1])

                    # 使用 is_close 函数来比较浮点数
                    points = [(ver_x1, ver_y1), (ver_x2, ver_y2), (ver_x3, ver_y3), (ver_x4, ver_y4)]
                    for point in points:
                        if is_close(point[0], min_x_cor, epsilon) or is_close(point[0], max_x_cor, epsilon) or (
                                min_x_cor < point[0] < max_x_cor):
                            if is_close(point[1], min_y_cor, epsilon) or is_close(point[1], max_y_cor, epsilon) or (
                                    min_y_cor < point[1] < max_y_cor):
                                if is_close(point[0], min_x_squa, epsilon) or is_close(point[0], max_x_squa,
                                                                                       epsilon) or (
                                        min_x_squa < point[0] < max_x_squa):
                                    if is_close(point[1], min_y_squa, epsilon) or is_close(point[1], max_y_squa,
                                                                                           epsilon) or (
                                            min_y_squa < point[1] < max_y_squa):
                                        cen_edge[i].append(edge)
                                        weight[i] += 1
                                        break

    weight_sort = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]
    max_cen_index = [j[0] for j in weight_sort]
    # print("NE: 删除前N个方格内的节点和边:", max_cen_index)
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_edge[pp]:
            remove_edges.append(q)
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)
    # net_copy.remove_nodes_from(list(nx.isolates(net_copy)))
    return net



def Calculate_CI(net, l=2):
    ci_dict = {}
    for i in net.nodes():
        ki = net.degree(i)
        ball_nodes = nx.single_source_shortest_path_length(net, i, cutoff=l)
        boundary_nodes = [node for node, dist in ball_nodes.items() if dist == l]
        ci = (ki - 1) * sum((net.degree(j) - 1) for j in boundary_nodes)
        ci_dict[i] = ci
    return ci_dict

def CIAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_dict = Calculate_CI(net, l=2)
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)

    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += cen_dict[j]
    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print("CI: 删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))  # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net



def PagerankAreas(net, grid, pos, num):

    node_list = nx.nodes(net)
    cen_node, weight = {}, {}       # 网格索引为键,网格内包含的节点的度和为值
    grid_index = list(grid.keys())  # 网格索引
    grid_pos = list(grid.values())  # 网格四个点坐标
    cen_dict = nx.pagerank(net)
    cen_edge, edge_weight = FindEdgesInGrid(g, Grid_with_index, node_coordinate)

    for i in grid_index:
        weight[i] = 0
        cen_node[i] = []
        for j in node_list:
            if grid_pos[i][0][0] <= pos[j][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= pos[j][1] <= grid_pos[i][2][1]:
                cen_node[i].append(j)
                weight[i] += cen_dict[j]
    sort_list2 = sorted(weight.items(), key=lambda it: it[1], reverse=True)
    sort_list = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]  # 按降序排列的元组列表
    max_cen_index = [j[0] for j in sort_list]     # 按权重降序排列的方格索引
    # print("PR: 删除前N个方格内的节点和边:", max_cen_index)
    remove_nodes = []
    remove_edges = []
    for pp in max_cen_index:
        for q in cen_node[pp]:  # 遍历该格点中包含的节点索引列表
            remove_nodes.append(q)
        for r in cen_edge[pp]:
            remove_edges.append(r)
    remove_nodes = list(set(remove_nodes))  # 去除重复的节点索引
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)

    return net



def EdgeBetweenness(net, Grid, pos, num):
    edge_list = list(net.edges)
    cen_edge, weight = {}, {}
    grid_index = list(Grid.keys())
    # print("网格索引：", grid_index)
    grid_pos = list(Grid.values())
    # print("网格四个点坐标：", grid_pos)
    edge_dict = nx.edge_betweenness_centrality(net)
    # print("边介数：", edge_dict)
    for i in grid_index:
        weight[i] = 0
        cen_edge[i] = []
        for edge in edge_list:
            cor = [pos[edge[0]], pos[edge[1]]]
            if (grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]) or \
               (grid_pos[i][0][0] <= cor[1][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[1][1] <= grid_pos[i][2][1]):
                cen_edge[i].append(edge)
                weight[i] += edge_dict[edge]
            else:
                if cor[0][1] - cor[1][1] == 0:
                    if cor[0][0] == grid_pos[i][0][0] or cor[0][0] == grid_pos[i][1][0]:
                        minx = min(cor[0][1], cor[1][1])
                        maxx = max(cor[0][1], cor[1][1])
                        if minx < grid_pos[i][0][1] < grid_pos[i][2][1] < maxx:
                            cen_edge[i].append(edge)
                            weight[i] += edge_dict[edge]
                    else:
                        u1 = cor[0][0]
                        v1 = grid_pos[i][0][1]
                        u2 = cor[0][0]
                        v2 = grid_pos[i][2][1]
                        if (grid_pos[i][0][0] < u1 < grid_pos[i][1][0] or grid_pos[i][0][1] < v1 < grid_pos[i][2][1]) and (cor[0][0] < u1 < cor[1][0] or cor[0][1] < v1 < cor[1][1])  or \
                           (grid_pos[i][0][0] < u2 < grid_pos[i][1][0] or grid_pos[i][0][1] < v2 < grid_pos[i][2][1]) and (cor[0][0] < u2 < cor[1][0] or cor[0][1] < v2 < cor[1][1]):
                            cen_edge[i].append(edge)
                            weight[i] += edge_dict[edge]

                elif cor[1][0] - cor[0][0] == 0:
                    if cor[0][1] == grid_pos[i][0][1] or cor[0][1] == grid_pos[i][2][1]:
                        minx = min(cor[0][0], cor[1][0])
                        maxx = max(cor[0][0], cor[1][0])
                        if minx < grid_pos[i][0][0] < grid_pos[i][1][0] < maxx:
                            cen_edge[i].append(edge)
                            weight[i] += edge_dict[edge]
                    else:
                        u1 = grid_pos[i][0][0]
                        v1 = cor[0][1]
                        u2 = grid_pos[i][1][0]
                        v2 = cor[0][1]
                        if (grid_pos[i][0][0] < u1 < grid_pos[i][1][0] or grid_pos[i][0][1] < v1 < grid_pos[i][2][1]) and (cor[0][0] < u1 < cor[1][0] or cor[0][1] < v1 < cor[1][1])  or \
                           (grid_pos[i][0][0] < u2 < grid_pos[i][1][0] or grid_pos[i][0][1] < v2 < grid_pos[i][2][1]) and (cor[0][0] < u2 < cor[1][0] or cor[0][1] < v2 < cor[1][1]):
                            cen_edge[i].append(edge)
                            weight[i] += edge_dict[edge]
                else:
                    slope = (cor[1][1] - cor[0][1]) / (cor[1][0] - cor[0][0])  # 斜率
                    intercept = cor[0][1] - slope * cor[0][0]  # 截距

                    ver_x1 = grid_pos[i][0][0]
                    ver_y1 = slope * ver_x1 + intercept

                    ver_x2 = grid_pos[i][1][0]
                    ver_y2 = slope * ver_x2 + intercept

                    ver_y3 = grid_pos[i][0][1]
                    ver_x3 = (ver_y3 - intercept)/slope

                    ver_y4 = grid_pos[i][2][1]
                    ver_x4 = (ver_y4 - intercept)/slope

                    min_x_cor = min(cor[0][0], cor[1][0])
                    max_x_cor = max(cor[0][0], cor[1][0])
                    min_y_cor = min(cor[0][1], cor[1][1])
                    max_y_cor = max(cor[0][1], cor[1][1])

                    min_x_squa = min(grid_pos[i][0][0], grid_pos[i][1][0])
                    max_x_squa = max(grid_pos[i][0][0], grid_pos[i][1][0])
                    min_y_squa = min(grid_pos[i][0][1], grid_pos[i][2][1])
                    max_y_squa = max(grid_pos[i][0][1], grid_pos[i][2][1])


                    if ((min_x_cor < ver_x1 < max_x_cor or min_y_cor < ver_y1 < max_y_cor) and (min_x_squa < ver_x1 < max_x_squa or min_y_squa < ver_y1 < max_y_squa)) or \
                       ((min_x_cor < ver_x2 < max_x_cor or min_y_cor < ver_y2 < max_y_cor) and (min_x_squa < ver_x2 < max_x_squa or min_y_squa < ver_y2 < max_y_squa)) or \
                       ((min_x_cor < ver_x3 < max_x_cor or min_y_cor < ver_y3 < max_y_cor) and (min_x_squa < ver_x3 < max_x_squa or min_y_squa < ver_y3 < max_y_squa)) or \
                       ((min_x_cor < ver_x4 < max_x_cor or min_y_cor < ver_y4 < max_y_cor) and (min_x_squa < ver_x4 < max_x_squa or min_y_squa < ver_y4 < max_y_squa)):
                        cen_edge[i].append(edge)
                        weight[i] += edge_dict[edge]
                    else:
                        continue

    # sort_list1 = sorted(weight.items(), key=lambda it: it[1], reverse=True)
    # print("边和权重降序: ", sort_list1)
    weight_sort = sorted(weight.items(), key=lambda it: it[1], reverse=True)[:num]
    remove_edges = []
    max_cen_index = [j[0] for j in weight_sort]
    # print("删除前N个方格内的节点和边:", max_cen_index)
    for pp in max_cen_index:
        for q in cen_edge[pp]:
            remove_edges.append(q)
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)
    # net_copy.remove_nodes_from(list(nx.isolates(net_copy)))
    return net


if __name__=="__main__":

    start_time = time.time()  # 记录开始时间
    interval = 1 / 15
    Grid_with_index = CreateGrid(interval)
    num_grids = len(Grid_with_index) + 1
    # num_grids = 51


    # seed = 2
    # g = nx.barabasi_albert_graph(100, 2, seed=seed)
    # g = nx.watts_strogatz_graph(100, 4, 0.2, seed=seed)
    # g = nx.newman_watts_strogatz_graph(100, 2, 0.1, seed=seed)
    # g = nx.erdos_renyi_graph(n=100, p=0.1, seed=seed)
    # layout = nx.random_layout(g, seed=seed)
    # coords = np.array(list(layout.values()))
    # node_coordinate = (coords - coords.min(axis=0)) / (coords.max(axis=0) - coords.min(axis=0))
    # node_coordinate = {node: list(round(coord, 4) for coord in layout[node]) for node in g.nodes()}  # 获取每个节点的坐标，四舍五入保留四位小数



    # 加载真实网络
    file1 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\1-Fiber.xlsx'
    file2 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Center.xlsx'  # 太大
    file3 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Friedrichshain.xlsx'
    file4 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Mitte-Center.xlsx'
    # file5 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Mitte-pre.xlsx'  # 完全没效果,放弃了
    file6 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Prenzlauerberg-Center.xlsx'
    file7 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Tiergarten.xlsx'
    # file8 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\3-Birmingham-England.xlsx'
    file9 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\4-ChicagoSketch.xlsx'  #
    file10 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\4-ChicagoRegional.xlsx'
    # file11 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\5-Goldcoast.xlsx'
    # file12 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\6-Philadelphia.xlsx'  # 大
    file13 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\7-SiouxFalls.xlsx'
    # file14 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\8-Sydney.xlsx'
    # file15 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\9-USAir.xlsx'
    # file16 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\10-Airlines.xlsx'
    file17 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\11-MinnesotaRoad.xlsx'  # 无
    # file18 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\12-Rail.xlsx'
    # file19 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\13-Central-chilean-power-grid.xlsx'
    file20 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\14-Coach.xlsx'
    file21 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\15-OldenburgRoad.xlsx'
    file22 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\16-CaliforniaRoad.xlsx'
    # file23 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\17-SanJoaquinCountyRoad.xlsx'
    # file24 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\18-co_Manchester.xlsx'

    # file25 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\19-geneve.xlsx'
    file26 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\20-ojohannesburg.xlsx'
    # file27 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\21-frankfurt.xlsx'
    # file28 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\22-sanfrancisco.xlsx'
    # file29 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\23-wellington.xlsx'
    file30 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\24-kigali.xlsx'
    file31= r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\25-casablanca.xlsx'
    file32= r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\26-hongkong.xlsx'

    file33 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\27-bogota.xlsx'
    file34 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\28-losangeles.xlsx'
    # file35 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\29-saopaulo.xlsx'
    file36 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\30-singapore.xlsx'
    # file37 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\31-newyork.xlsx'
    file38 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\32-riodejaneiro.xlsx'
    file39 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\33-santiago.xlsx'
    # file40 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\34-shanghai.xlsx'
    file41 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\35-london.xlsx'
    file42 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\36-melbourne.xlsx'



    files = [file1, file2, file3, file4, file6, file7, file9, file10, file13, file17, file20, file21, file22,  file26, file30, file31, file32, file33, file34, file36, file38, file39, file41, file42]
    i = 1
    for file in files:

        print("程序运行到第{}个文件，请勿关闭！！！".format(i))

        g, node_coordinate = load_network(file)
        LCC_Degree = []
        LCC_Eigenvector = []
        LCC_Numedges = []
        LCC_CI = []
        LCC_PageRank = []
        LCC_AverageDegree = []

        for Num in range(1, num_grids):
            G_copy1, G_copy2, G_copy3, G_copy4, G_copy5, G_copy6 = copy.deepcopy(g), copy.deepcopy(g), copy.deepcopy(g), copy.deepcopy(g), copy.deepcopy(g), copy.deepcopy(g)

            # LCCNet1 = DegreeAreas(G_copy1, Grid_with_index, node_coordinate, Num)
            # LCCNet2 = EigenvectorAreas(G_copy2, Grid_with_index, node_coordinate, Num)
            # LCCNet3 = NumEdges(G_copy3, Grid_with_index, node_coordinate, Num)
            # LCCNet4 = CIAreas(G_copy4, Grid_with_index, node_coordinate, Num)
            # LCCNet5 = PagerankAreas(G_copy5, Grid_with_index, node_coordinate, Num)
            LCCNet6 = AverageDegreeAreas(G_copy6, Grid_with_index, node_coordinate, Num)

            # Lcc1 = PerformanceEvaluation(LCCNet1)
            # Lcc2 = PerformanceEvaluation(LCCNet2)
            # Lcc3 = PerformanceEvaluation(LCCNet3)
            # Lcc4 = PerformanceEvaluation(LCCNet4)
            # Lcc5 = PerformanceEvaluation(LCCNet5)
            Lcc6 = PerformanceEvaluation(LCCNet6)

            # LCC_Degree.append(Lcc1)
            # LCC_Eigenvector.append(Lcc2)
            # LCC_Numedges.append(Lcc3)
            # LCC_CI.append(Lcc4)
            # LCC_PageRank.append(Lcc5)
            LCC_AverageDegree.append(Lcc6)


        # 将列表转换为 DataFrame
        sheetname = 'LCC'
        # df = pd.DataFrame({'Degree': LCC_Degree, 'Eigenvector': LCC_Eigenvector, 'NumEdges': LCC_Numedges, 'CI': LCC_CI, 'PageRank': LCC_PageRank})
        df = pd.DataFrame({'AverageDegree': LCC_AverageDegree})
        # df.to_excel(r"D:\Python_Projects\Paper2-Network-Disintegration\Results\GraphSAGE-12345\100-10-Airlines.xlsx", sheet_name=sheetname, index=False)

        # 获取文件名（不包括扩展名）
        base_name = os.path.basename(file)
        file_name, _ = os.path.splitext(base_name)

        # 使用文件名作为表名和 Excel 文件名
        # sheetname = file_name
        # excel_file = f"D:\\Python_Projects\\Paper2-Network-Disintegration\\Results\\900\\{file_name}.xlsx"
        excel_file = f"D:\\Python_Projects\\Paper2-Network-Disintegration\\Results\\Supplementary_experiments\\{file_name}.xlsx"
        df.to_excel(excel_file, sheet_name=sheetname, index=False)
        i = i+1
    end_time = time.time()  # 记录结束时间
    print("程序运行时间：", end_time - start_time, "秒")

