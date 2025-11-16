import torch
import networkx as nx
import openpyxl as xl
import numpy as np
from GenerateTrainDataSet import *
from Model import *
from CentralityMenthod import *
import copy
import openpyxl as xl
import matplotlib.pyplot as plt
from torch_geometric.data import Data
import pandas as pd
import time
from scipy.io import savemat


def load_network(file):

    wb = xl.load_workbook(file)
    ws_node, ws_link = wb['node'], wb['link']
    row_link, row_node = ws_link.max_row, ws_node.max_row
    edge_list = list(ws_link.iter_rows(min_col=1, max_col=2, min_row=2, max_row=row_link, values_only=True))
    node_list = list(ws_node.iter_rows(min_col=1, max_col=3, min_row=2, max_row=row_node, values_only=True))
    G = nx.Graph()
    for i in node_list:
        G.add_node(i[0])
    for i in edge_list:
        G.add_edge(i[0], i[1])

    x_all, y_all = [i[1] for i in node_list], [i[2] for i in node_list]
    xcor_min, xcor_max = min(x_all), max(x_all)
    xcor = xcor_max - xcor_min
    ycor_min, ycor_max = min(y_all), max(y_all)
    ycor = ycor_max - ycor_min
    node_position = {}  # 节点的归一化位置坐标
    for i in node_list:
        node_position[i[0]] = list(
            (np.array([i[1], i[2]]) - np.array([xcor_min, ycor_min])) / np.array([xcor, ycor]))  # 计算每个节点的归一化位置

    return G, node_position



def CreateGrid(interval):
    """生成网格字典,键为网格索引，值为网格四个点的坐标"""
    part = int(1 / interval)  # 生成一个 part × part 的二维网格
    grid = [[[(round(interval * m, 4), round(interval * n, 4)),
              (round(interval * (m + 1), 4), round(interval * n, 4)),
              (round(interval * m, 4), round(interval * (n + 1), 4)),
              (round(interval * (m + 1), 4), round(interval * (n + 1), 4))]
              for n in range(part)] for m in range(part)]

    # 将二维网格转换为一维列表
    Grid_flat = [coord for sublist in grid for coord in sublist]
    # 给每个网格单元添加一个索引
    Grid = {i: Grid_flat[i] for i in range(len(Grid_flat))}

    return Grid


def Get_Grid_Adj(Grid, cen_edge):
    size = len(Grid)
    Grid_adj = np.zeros((size, size))
    for i in range(0, size):
        for j in range(0, size):
            if i != j and len(set(cen_edge[i]).intersection(cen_edge[j])) > 0:
                Grid_adj[i][j] = len(set(cen_edge[i]).intersection(cen_edge[j]))
                Grid_adj[j][i] = len(set(cen_edge[i]).intersection(cen_edge[j]))
    return Grid_adj

def Get_Grid_Adj1(Grid, cen_edge):
    size = len(Grid)
    Grid_adj = np.zeros((size, size))
    for i in range(0, size):
        for j in range(0, size):
            if i != j and len(set(cen_edge[i]).intersection(cen_edge[j])) > 0:
                Grid_adj[i][j] = 1
                Grid_adj[j][i] = 1
    return Grid_adj


epsilon = 1e-9  # 定义一个小的epsilon值
def is_close(a, b, epsilon):
    """比较两个浮点数是否在epsilon范围内相等"""
    return abs(a - b) < epsilon

def FindEdgesInGrid(net, Grid, pos):
    edge_list = list(net.edges)
    cen_edge, weight = {}, {}
    grid_index = list(Grid.keys()) # 网格索引
    grid_pos = list(Grid.values()) # 网格四个点坐标
    for i in grid_index:
        weight[i] = 0
        cen_edge[i] = []
        for edge in edge_list:
            cor = [pos[edge[0]], pos[edge[1]]]  # 判断至少有一个端点在方格内
            if (grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]) or \
               (grid_pos[i][0][0] <= cor[1][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[1][1] <= grid_pos[i][2][1]):
                cen_edge[i].append(edge)
                weight[i] += 1
            else:  # 两个端点都不在方格内
                if cor[0][0] - cor[1][0] == 0:  # 平行y轴

                    if grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0]:
                        miny = min(cor[0][1], cor[1][1])
                        maxy = max(cor[0][1], cor[1][1])
                        if miny <= grid_pos[i][0][1] and maxy >= grid_pos[i][2][1]:
                            cen_edge[i].append(edge)
                            weight[i] += 1

                elif cor[0][1] - cor[1][1] == 0:  # 平行x轴

                    if grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]:
                        minx = min(cor[0][0], cor[1][0])
                        maxx = max(cor[0][0], cor[1][0])
                        if minx <= grid_pos[i][0][0] and maxx >= grid_pos[i][1][0]:
                            cen_edge[i].append(edge)
                            weight[i] += 1

                else:
                    slope = (cor[1][1] - cor[0][1]) / (cor[1][0] - cor[0][0])  # 斜率
                    intercept = cor[0][1] - slope * cor[0][0]  # 截距

                    ver_x1 = grid_pos[i][0][0]
                    ver_y1 = slope * ver_x1 + intercept
                    ver_x2 = grid_pos[i][1][0]
                    ver_y2 = slope * ver_x2 + intercept
                    ver_y3 = grid_pos[i][0][1]
                    ver_x3 = (ver_y3 - intercept)/slope
                    ver_y4 = grid_pos[i][2][1]
                    ver_x4 = (ver_y4 - intercept)/slope

                    min_x_cor = min(cor[0][0], cor[1][0])
                    max_x_cor = max(cor[0][0], cor[1][0])
                    min_y_cor = min(cor[0][1], cor[1][1])
                    max_y_cor = max(cor[0][1], cor[1][1])

                    min_x_squa = min(grid_pos[i][0][0], grid_pos[i][1][0])
                    max_x_squa = max(grid_pos[i][0][0], grid_pos[i][1][0])
                    min_y_squa = min(grid_pos[i][0][1], grid_pos[i][2][1])
                    max_y_squa = max(grid_pos[i][0][1], grid_pos[i][2][1])

                    # 使用 is_close 函数来比较浮点数
                    points = [(ver_x1, ver_y1), (ver_x2, ver_y2), (ver_x3, ver_y3), (ver_x4, ver_y4)]
                    for point in points:
                        if is_close(point[0], min_x_cor, epsilon) or is_close(point[0], max_x_cor, epsilon) or (min_x_cor < point[0] < max_x_cor):
                            if is_close(point[1], min_y_cor, epsilon) or is_close(point[1], max_y_cor, epsilon) or (min_y_cor < point[1] < max_y_cor):
                                if is_close(point[0], min_x_squa, epsilon) or is_close(point[0], max_x_squa,epsilon) or (min_x_squa < point[0] < max_x_squa):
                                    if is_close(point[1], min_y_squa, epsilon) or is_close(point[1], max_y_squa,epsilon) or (min_y_squa < point[1] < max_y_squa):
                                        cen_edge[i].append(edge)
                                        weight[i] += 1
                                        break

    return cen_edge, weight



def PerformanceEvaluation(net):
    """计算最大连通子图比例"""
    size = len(nx.nodes(net))
    largest_cc = max(nx.connected_components(net), key=len)  # 获取最大的连通子图
    largest_cc = len(largest_cc) / size
    return largest_cc

def Machine(net, Grid, pos,max_cen_index, num):
    edge_list = list(net.edges)
    cen_edge, weight = {}, {}
    grid_index = list(Grid.keys())
    # print("网格索引：", grid_index)
    grid_pos = list(Grid.values())
    # print("网格四个点坐标：", grid_pos)
    for i in grid_index:
        weight[i] = 0
        cen_edge[i] = []
        for edge in edge_list:
            cor = [pos[edge[0]], pos[edge[1]]]  # 判断至少有一个端点在方格内
            if (grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[0][1] <=
                grid_pos[i][2][1]) or \
                    (grid_pos[i][0][0] <= cor[1][0] <= grid_pos[i][1][0] and grid_pos[i][0][1] <= cor[1][1] <=
                     grid_pos[i][2][1]):
                cen_edge[i].append(edge)
                weight[i] += 1
            else:  # 两个端点都不在方格内
                if cor[0][0] - cor[1][0] == 0:  # 平行y轴

                    if grid_pos[i][0][0] <= cor[0][0] <= grid_pos[i][1][0]:
                        # 检查网格单元格的y坐标是否在线段的y坐标范围内
                        miny = min(cor[0][1], cor[1][1])
                        maxy = max(cor[0][1], cor[1][1])
                        if miny <= grid_pos[i][0][1] and maxy >= grid_pos[i][2][1]:
                            cen_edge[i].append(edge)
                            weight[i] += 1



                elif cor[0][1] - cor[1][1] == 0:  # 平行x轴

                    if grid_pos[i][0][1] <= cor[0][1] <= grid_pos[i][2][1]:
                        # 检查网格单元格的y坐标是否在线段的y坐标范围内
                        minx = min(cor[0][0], cor[1][0])
                        maxx = max(cor[0][0], cor[1][0])
                        if minx <= grid_pos[i][0][0] and maxx >= grid_pos[i][1][0]:
                            cen_edge[i].append(edge)
                            weight[i] += 1

                else:

                    slope = (cor[1][1] - cor[0][1]) / (cor[1][0] - cor[0][0])  # 斜率
                    intercept = cor[0][1] - slope * cor[0][0]  # 截距
                    # print(slope)
                    # print(intercept)

                    ver_x1 = grid_pos[i][0][0]
                    ver_y1 = slope * ver_x1 + intercept

                    ver_x2 = grid_pos[i][1][0]
                    ver_y2 = slope * ver_x2 + intercept

                    ver_y3 = grid_pos[i][0][1]
                    ver_x3 = (ver_y3 - intercept) / slope

                    ver_y4 = grid_pos[i][2][1]
                    ver_x4 = (ver_y4 - intercept) / slope

                    min_x_cor = min(cor[0][0], cor[1][0])
                    max_x_cor = max(cor[0][0], cor[1][0])
                    min_y_cor = min(cor[0][1], cor[1][1])
                    max_y_cor = max(cor[0][1], cor[1][1])

                    min_x_squa = min(grid_pos[i][0][0], grid_pos[i][1][0])
                    max_x_squa = max(grid_pos[i][0][0], grid_pos[i][1][0])
                    min_y_squa = min(grid_pos[i][0][1], grid_pos[i][2][1])
                    max_y_squa = max(grid_pos[i][0][1], grid_pos[i][2][1])

                    # 使用 is_close 函数来比较浮点数
                    points = [(ver_x1, ver_y1), (ver_x2, ver_y2), (ver_x3, ver_y3), (ver_x4, ver_y4)]
                    for point in points:
                        if is_close(point[0], min_x_cor, epsilon) or is_close(point[0], max_x_cor, epsilon) or (
                                min_x_cor < point[0] < max_x_cor):
                            if is_close(point[1], min_y_cor, epsilon) or is_close(point[1], max_y_cor, epsilon) or (
                                    min_y_cor < point[1] < max_y_cor):
                                if is_close(point[0], min_x_squa, epsilon) or is_close(point[0], max_x_squa,
                                                                                       epsilon) or (
                                        min_x_squa < point[0] < max_x_squa):
                                    if is_close(point[1], min_y_squa, epsilon) or is_close(point[1], max_y_squa,
                                                                                           epsilon) or (
                                            min_y_squa < point[1] < max_y_squa):
                                        cen_edge[i].append(edge)
                                        weight[i] += 1
                                        break

    remove_edges = []

    # max_cen_index = max_cen_index[:num]
    max_cen_index = [j for j in max_cen_index[:num]]

    for pp in max_cen_index:
        for q in cen_edge[pp]:
            remove_edges.append(q)
    remove_edges = list(set(remove_edges))
    net.remove_edges_from(remove_edges)
    # net_copy.remove_nodes_from(list(nx.isolates(net_copy)))
    return net




if __name__=="__main__":

    interval = 1 / 10
    Grid = CreateGrid(interval)
    num_grids = len(Grid) + 1
    # num_grids = 50


    # 加载合成网络
    # seed = 2
    # # g = nx.barabasi_albert_graph(100, 2, seed=seed)
    # # g = nx.watts_strogatz_graph(100, 4, 0.2, seed=seed)
    # # g = nx.newman_watts_strogatz_graph(100, 2, 0.1, seed=seed)
    # g = nx.erdos_renyi_graph(n=100, p=0.1, seed=seed)
    # layout = nx.random_layout(g, seed=seed)
    # # coords = np.array(list(layout.values()))
    # # node_coordinate = (coords - coords.min(axis=0)) / (coords.max(axis=0) - coords.min(axis=0))
    # node_coordinate = {node: list(round(coord, 4) for coord in layout[node]) for node in g.nodes()}  # 获取每个节点的坐标，四舍五入保留四位小数


    # 加载真实网络

    file1 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\1-Fiber.xlsx'
    file2 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Center.xlsx'# 太大
    file3 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Friedrichshain.xlsx'
    file4 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Mitte-Center.xlsx'
    file5 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Mitte-pre.xlsx' # 完全没效果,放弃了
    file6 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Prenzlauerberg-Center.xlsx'
    file7 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\2-Berlin-Tiergarten.xlsx'
    file8 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\3-Birmingham-England.xlsx'
    file9 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\4-ChicagoSketch.xlsx'  #
    file10 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\4-ChicagoRegional.xlsx'
    file11 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\5-Goldcoast.xlsx'
    file12 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\6-Philadelphia.xlsx' # 大
    file13 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\7-SiouxFalls.xlsx'
    file14 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\8-Sydney.xlsx'
    file15 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\9-USAir.xlsx'
    file16 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\10-Airlines.xlsx'
    file17 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\11-MinnesotaRoad.xlsx' # 有
    file18 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\12-Rail.xlsx'
    file19 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\13-Central-chilean-power-grid.xlsx'
    file20 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\14-Coach.xlsx'
    file21 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\15-OldenburgRoad.xlsx'
    file22 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\16-CaliforniaRoad.xlsx'
    file23 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\17-SanJoaquinCountyRoad.xlsx'
    file24 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\18-co_Manchester.xlsx'

    file25 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\19-geneve.xlsx'
    file26 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\20-ojohannesburg.xlsx'
    file27 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\21-frankfurt.xlsx'
    file28 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\22-sanfrancisco.xlsx'
    file29 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\23-wellington.xlsx'
    file30 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\24-kigali.xlsx'
    file31= r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\25-casablanca.xlsx'
    file32= r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\26-hongkong.xlsx'

    file33 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\27-bogota.xlsx'
    file34 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\28-losangeles.xlsx'
    file35 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\29-saopaulo.xlsx'
    file36 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\30-singapore.xlsx'
    file37 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\31-newyork.xlsx'
    file38 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\32-riodejaneiro.xlsx'
    file39 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\33-santiago.xlsx'
    file40 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\34-shanghai.xlsx'
    file41 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\35-london.xlsx'
    file42 = r'D:\Python_Projects\Paper2-Network-Disintegration\Real-World-Networks\36-melbourne.xlsx'

    files = [file1]
    # files = [file25, file26, file27, file28, file29, file30, file31, file32, file33, file34, file35, file36, file37, file38, file39, file40, file41, file42]
    # files = [file1, file2, file3, file4, file5, file6, file7, file8, file9, file10, file11, file12, file13, file14, file15, file16, file17, file18, file19, file20,file21, file22, file23, file24, file25, file26, file27, file28, file29, file30, file31, file32, file33, file34, file35, file36, file37, file38, file39, file40, file41, file42]
    All_LCC = []

    for file in files:

        start_time = time.time()  # 记录开始时间

        g, node_coordinate = load_network(file)
        cen_edge, weight = FindEdgesInGrid(g, Grid, node_coordinate)

        Grid_adj_matrix = np.array(Get_Grid_Adj(Grid, cen_edge))
        print(np.sum(Grid_adj_matrix))
        Grid_adj_matrix1 = np.array(Get_Grid_Adj1(Grid, cen_edge))
        print(np.sum(Grid_adj_matrix1))
        # 创建一个字典，将矩阵包含在其中
        data_dict = {'Fiber_adj': Grid_adj_matrix1}
        # 将数据字典保存为MATLAB兼容的.mat文件
        savemat('F:\MATLAB_Code\Fiber_adj.mat', data_dict)

        min_val = np.min(Grid_adj_matrix)
        max_val = np.max(Grid_adj_matrix)
        Grid_adj_matrix = (Grid_adj_matrix - min_val) / (max_val - min_val)


        source_nodes, target_nodes = np.nonzero(Grid_adj_matrix[:, :])  # 获取邻接矩阵中非零元素的索引，即图中存在边的节点对

        edge_weight = Grid_adj_matrix[source_nodes, target_nodes]
        edge_weight = torch.tensor(edge_weight, dtype=torch.float)
        # print(edge_weight.shape)
        # edge_weight = edge_weight
        # print(edge_weight.shape)

        source_nodes = source_nodes.reshape((1, -1))
        target_nodes = target_nodes.reshape((1, -1))
        edge_index = torch.tensor(np.concatenate((source_nodes, target_nodes), axis=0), dtype=torch.long)


        # print("edgeIndex", edge_index)

        Grid_features_matrix = Generate_Grid_Feature(g, Grid, node_coordinate)
        x = torch.tensor(Grid_features_matrix[:, :], dtype=torch.float)

        data = Data(x=x, edge_index=edge_index.view(2, -1), edge_weight=edge_weight) # edge_index.view(2, -1)
        # print(data.x.shape)
        # print(data.edge_index.shape)
        # print(data.edge_weight.shape)
        model = SNDM()

        device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

        train_well_mode = r'D:\Python_Projects\Paper2-Network-Disintegration\Checkpoints\best_model.pkl'
        checkpoint = torch.load(train_well_mode, map_location=device)
        model.load_state_dict(checkpoint)
        model.eval()  # 模型设置为评估模式



        predictions = model(data.x, data.edge_index, data.edge_weight)
        # predictions = model(data.x)
        # print(predictions.shape)
        # print(len(predictions))
        # 将预测值转换为列表
        predictions_list = predictions.tolist()
        # print(predictions_list)

        # 创建一个索引列表
        indices = list(range(len(predictions_list)))

        # 将索引和预测值打包并排序
        sorted_pairs = sorted(zip(predictions_list, indices), reverse=True)

        # 解压排序后的预测值和索引
        sorted_predictions, sorted_indices = zip(*sorted_pairs)
        sorted_indices = list(sorted_indices)
        print("排序后的预测值：", sorted_predictions)
        print("对应的索引：", sorted_indices)


        # plt.figure(figsize=(6, 6))
        # plt.title("Spatial Network")
        # # plt.xlabel("X")
        # # plt.ylabel("Y")
        # plt.xlim(-0.05, 1.05)
        # plt.ylim(-0.05, 1.05)
        # plt.xticks(np.arange(0, 1.1, 0.05))
        # plt.yticks(np.arange(0, 1.1, 0.05))
        # plt.grid(True)
        # # layout = nx.random_layout(g)  # 随机布局
        # # node_coordinate = {node: list(round(coord, 4) for coord in layout[node]) for node in G.nodes()} # 获取每个节点的坐标，四舍五入保留四位小数
        # # 嵌入节点到网格中
        # for node, position in node_coordinate.items():
        #     x = position[0]
        #     y = position[1]
        #     plt.scatter(x, y, color='blue', s=10)  # 绘制节点
        #     for neighbor in g.neighbors(node):
        #         neighbor_pos = node_coordinate[neighbor]
        #         plt.plot([x, neighbor_pos[0]], [y, neighbor_pos[1]], color='g', linewidth=0.1)
        #     # 可以根据需要添加节点标签
        #     # plt.text(x, y, str(node), fontsize=12, ha='right', va='bottom')
        # plt.show()

        LCC_Machine = []
        for Num in range(1, num_grids):

            G_copy=copy.deepcopy(g)
            LCCNet = Machine(G_copy, Grid, node_coordinate, sorted_indices, Num)
            Lcc = PerformanceEvaluation(LCCNet)
            LCC_Machine.append(Lcc)


        end_time = time.time()  # 记录结束时间
        print("程序运行时间：", end_time - start_time, "秒")
        All_LCC.append(LCC_Machine)

    # 初始化一个空的 DataFrame
    df = pd.DataFrame()

    # 循环遍历 All_LCC 列表
    for file, LCC_Machine in zip(files, All_LCC):
        # 获取文件名
        filename = file.split('\\')[-1].split('.')[0]

        # 将每个 LCC_Machine 列表转换为 Series 对象，并添加到 DataFrame 中
        df[filename] = pd.Series(LCC_Machine)

    # 保存到 Excel
    # df.to_excel(r"D:\Python_Projects\Paper2-Network-Disintegration\Results\results.xlsx",
    #             sheet_name='LCC', index=False)



















